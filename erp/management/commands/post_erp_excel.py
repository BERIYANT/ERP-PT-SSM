from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from erp.models import (
    BusinessPartner,
    CashTransaction,
    CustomerPurchaseOrder,
    DocumentStatus,
    ImportBatch,
    Invoice,
    InvoiceItem,
    Payment,
    Project,
    ProjectSegment,
    PurchaseOrderItem,
    UserOrganization,
)
from portal.models import User


PO_FILE = "FIBERISASI PO KOPINDOSAT BB ROLL JALUR UTARA.xlsx"
INVOICE_FILE = "INVOICE KOPINDOSAT.xlsx"


def money(value):
    if value in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError, TypeError):
        return Decimal("0")


def quantity(value):
    if value in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(value)).quantize(Decimal("0.0001"))
    except (InvalidOperation, ValueError, TypeError):
        return Decimal("0")


def normalize_installments(contract_value, installments):
    """Distribute the PO value over its recorded payment installments."""
    valid = [item for item in installments if item[0] and item[1] > 0 and item[2]]
    source_total = sum((item[1] for item in valid), Decimal("0"))
    if contract_value <= 0 or source_total <= 0:
        return []
    normalized = []
    allocated = Decimal("0")
    for index, (number, amount, paid_date, sequence) in enumerate(valid):
        if index == len(valid) - 1:
            normalized_amount = contract_value - allocated
        else:
            normalized_amount = (contract_value * amount / source_total).quantize(Decimal("0.01"))
            allocated += normalized_amount
        normalized.append((number, normalized_amount, paid_date, sequence))
    return normalized


def excel_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return (datetime(1899, 12, 30) + timedelta(days=value)).date()
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value).date()
        except ValueError:
            return None
    return None


class Command(BaseCommand):
    help = "Posting PO, proyek, invoice, pembayaran, dan cash-in Excel tervalidasi ke transaksi ERP."

    def add_arguments(self, parser):
        parser.add_argument("--directory", default="/Users/mac/Downloads/Excel SSM")
        parser.add_argument("--username", default="superadmin")

    @transaction.atomic
    def handle(self, *args, **options):
        directory = Path(options["directory"])
        po_path, invoice_path = directory / PO_FILE, directory / INVOICE_FILE
        if not po_path.exists() or not invoice_path.exists():
            raise CommandError("Workbook PO atau invoice tidak ditemukan.")
        user = User.objects.filter(username=options["username"]).first()
        organization = UserOrganization.objects.filter(user=user).select_related("company").first()
        if not user or not organization:
            raise CommandError("User import belum terhubung ke perusahaan ERP.")
        company = organization.company
        client, _ = BusinessPartner.objects.get_or_create(
            company=company,
            partner_type=BusinessPartner.PartnerType.CLIENT,
            name="KOPINDOSAT",
        )
        summary = {"projects": 0, "segments": 0, "purchase_orders": 0, "po_items": 0, "invoices": 0, "invoice_items": 0, "payments": 0, "cash_in": 0, "skipped": []}
        po_map = self._post_po(po_path, company, client, user, summary)
        self._post_invoices(invoice_path, po_map, company, client, user, summary)
        for file_name in (PO_FILE, INVOICE_FILE):
            ImportBatch.objects.filter(file_name=file_name).update(status=ImportBatch.Status.POSTED)
        self.stdout.write(self.style.SUCCESS(f"Posting Excel selesai: {summary}"))

    def _project_and_po(self, company, client, user, po_number, description, po_date, amount, summary):
        project, made = Project.objects.get_or_create(
            company=company,
            project_code=f"EXCEL-{po_number}",
            defaults={"client": client, "name": description, "start_date": po_date, "status": "ACTIVE"},
        )
        summary["projects"] += int(made)
        if project.name != description and description:
            project.name = description
            project.save(update_fields=["name", "updated_at"])
        segment, made = ProjectSegment.objects.get_or_create(
            project=project,
            segment_code="MAIN",
            defaults={"segment_name": description, "location": description},
        )
        summary["segments"] += int(made)
        po, made = CustomerPurchaseOrder.objects.get_or_create(
            project=project,
            po_number=po_number,
            defaults={"po_date": po_date or date.today(), "contract_value": amount, "status": DocumentStatus.APPROVED, "created_by": user},
        )
        summary["purchase_orders"] += int(made)
        changed = False
        if amount and po.contract_value != amount:
            po.contract_value, changed = amount, True
        if po_date and po.po_date != po_date:
            po.po_date, changed = po_date, True
        if changed:
            po.save(update_fields=["contract_value", "po_date", "updated_at"])
        return project, segment, po

    def _post_po(self, path, company, client, user, summary):
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = workbook["BB"]
        po_map = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            po_number = str(row[1] or "").strip()
            description = str(row[3] or "").strip()
            amount = money(row[4])
            if not po_number.startswith("PO") or not description or amount <= 0:
                continue
            project, segment, po = self._project_and_po(company, client, user, po_number, description, excel_date(row[2]), amount, summary)
            item, made = PurchaseOrderItem.objects.get_or_create(
                purchase_order=po,
                item_code="CONTRACT",
                defaults={"description": description, "unit": "ls", "qty": Decimal("1"), "unit_price": amount},
            )
            summary["po_items"] += int(made)
            po_map[po_number] = (project, segment, po)
            payment_columns = []
            for invoice_number, paid, paid_date, sequence in ((row[5], row[6], row[7], "1"), (row[8], row[9], row[10], "2")):
                invoice_number = str(invoice_number or "").strip()
                paid_amount = money(paid)
                payment_date = excel_date(paid_date)
                if not invoice_number or paid_amount <= 0 or not payment_date or "****" in str(paid):
                    continue
                payment_columns.append((invoice_number, paid_amount, payment_date, sequence))
            for invoice_number, paid_amount, payment_date, sequence in normalize_installments(amount, payment_columns):
                invoice = Invoice.objects.filter(project=project, invoice_number=invoice_number).order_by("-revision").first()
                if not invoice:
                    invoice = Invoice.objects.create(project=project, purchase_order=po, invoice_number=invoice_number, invoice_date=payment_date, subtotal=paid_amount, tax=0, status=DocumentStatus.PAID, created_by=user)
                    summary["invoices"] += 1
                self._payment(invoice, payment_date, paid_amount, f"EXCEL-{po_number}-{sequence}", user, summary)
        return po_map

    def _post_invoices(self, path, po_map, company, client, user, summary):
        formulas = openpyxl.load_workbook(path, read_only=True, data_only=False)
        values = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet_name in values.sheetnames:
            lower = sheet_name.lower()
            if "cancel" in lower or "revisi" in lower:
                summary["skipped"].append(sheet_name)
                continue
            sheet, source = values[sheet_name], formulas[sheet_name]
            invoice_number = str(sheet["M2"].value or sheet_name).strip()
            po_number = str(sheet["C13"].value or "").strip()
            description = str(sheet["C14"].value or "").strip()
            span = str(sheet["C15"].value or description).strip()
            invoice_date = excel_date(sheet["M3"].value)
            if not invoice_number.startswith("PJ.FO.") or not po_number.startswith("PO") or not invoice_date:
                continue
            subtotal = money(sheet["M51"].value or sheet["M49"].value)
            tax = money(sheet["M53"].value)
            total = money(sheet["M54"].value) or subtotal + tax
            if total <= 0:
                summary["skipped"].append(sheet_name)
                continue
            if po_number in po_map:
                project, segment, po = po_map[po_number]
            else:
                project, segment, po = self._project_and_po(company, client, user, po_number, description or span, invoice_date, money(sheet["H49"].value), summary)
                po_map[po_number] = (project, segment, po)
            if span and segment.segment_name != span:
                segment.segment_name = span
                segment.location = span
                segment.save(update_fields=["segment_name", "location", "updated_at"])
            invoice, made = Invoice.objects.get_or_create(
                project=project,
                invoice_number=invoice_number,
                revision=1,
                defaults={"purchase_order": po, "invoice_date": invoice_date, "subtotal": subtotal, "tax": tax, "status": DocumentStatus.SENT, "created_by": user},
            )
            summary["invoices"] += int(made)
            if not made:
                invoice.purchase_order, invoice.invoice_date, invoice.subtotal, invoice.tax = po, invoice_date, subtotal, tax
                invoice.save()
            if invoice.items.exists():
                continue
            for row_number in range(19, 49):
                description = str(sheet.cell(row_number, 3).value or "").strip()
                qty = quantity(sheet.cell(row_number, 10).value)
                unit_price = money(sheet.cell(row_number, 11).value)
                amount = money(sheet.cell(row_number, 13).value)
                if not description or qty <= 0 or amount <= 0:
                    continue
                if unit_price <= 0:
                    unit_price = (amount / qty).quantize(Decimal("0.01"))
                InvoiceItem.objects.create(invoice=invoice, description=description, qty=qty, unit_price=unit_price, item_type="EXCEL_ACTUAL")
                summary["invoice_items"] += 1

    def _payment(self, invoice, payment_date, amount, reference, user, summary):
        payment, made = Payment.objects.get_or_create(
            reference_number=reference,
            defaults={"invoice": invoice, "payment_date": payment_date, "amount": amount, "method": "BANK", "created_by": user},
        )
        summary["payments"] += int(made)
        cash, made = CashTransaction.objects.get_or_create(
            reference_number=f"CASH-{reference}",
            defaults={"project": invoice.project, "transaction_date": payment_date, "direction": CashTransaction.Direction.IN, "category": "CUSTOMER_PAYMENT", "amount": amount, "source_type": "PAYMENT", "source_id": payment.id},
        )
        summary["cash_in"] += int(made)
        paid = sum((p.amount for p in invoice.payments.all()), Decimal("0"))
        invoice.status = DocumentStatus.PAID if paid >= invoice.total else DocumentStatus.PARTIALLY_PAID
        invoice.save(update_fields=["status", "updated_at"])
