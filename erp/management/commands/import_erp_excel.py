import hashlib
import json
import re
from datetime import datetime
from pathlib import Path

import openpyxl
import xlrd
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from erp.models import ImportBatch, ImportRow
from portal.models import User


EXPECTED_FILES = {
    "Cash Flow BB Sby Gubeng - Wonokromo - Malang - Jombang.xlsx": "CASH_FLOW",
    "Expense Report Lap.xlsx": "EXPENSE",
    "FIBERISASI PO KOPINDOSAT BB ROLL JALUR UTARA.xlsx": "PURCHASE_ORDER",
    "Form Cash Flow Project.xls": "FUND_AND_CASH_FORM",
    "Form Pengajuan Dana.xlsx": "FUND_FORM",
    "INVOICE KOPINDOSAT.xlsx": "INVOICE",
    "Kopindosat RollOut Utara.xlsx": "PROJECT_ROLLOUT",
}


def canonical(value):
    if value is None:
        return ""
    if isinstance(value, (datetime,)):
        return value.isoformat()
    return re.sub(r"\s+", " ", str(value)).strip()


def fingerprint(values):
    payload = json.dumps([canonical(value) for value in values], ensure_ascii=False, separators=(",", ":"))
    return hashlib.sha256(payload.encode()).hexdigest()


def row_status(sheet_name, values):
    text = " ".join(canonical(value) for value in values).lower()
    errors = []
    if "****" in text:
        errors.append("Formula Excel rusak dan tidak boleh ditebak.")
    if "cancel" in sheet_name.lower() or "batal" in sheet_name.lower():
        errors.append("Dokumen sumber berstatus batal/cancel.")
    if "revisi" in sheet_name.lower():
        errors.append("Dokumen merupakan revisi dan memerlukan resolusi versi.")
    return (ImportRow.Status.INVALID if any("rusak" in error for error in errors) else ImportRow.Status.CONFLICT if errors else ImportRow.Status.VALID), errors


class Command(BaseCommand):
    help = "Membaca tujuh workbook ERP ke staging secara idempoten tanpa langsung memposting transaksi."

    def add_arguments(self, parser):
        parser.add_argument("--directory", default="/Users/mac/Downloads/Excel SSM")
        parser.add_argument("--username", default="superadmin")

    @transaction.atomic
    def handle(self, *args, **options):
        directory = Path(options["directory"])
        if not directory.is_dir():
            raise CommandError(f"Direktori tidak ditemukan: {directory}")
        user = User.objects.filter(username=options["username"]).first()
        if not user:
            raise CommandError("Jalankan seed_erp atau tentukan --username yang valid.")
        summary = {"files": 0, "sheets": 0, "rows": 0, "skipped_files": []}
        for file_name, source_type in EXPECTED_FILES.items():
            path = directory / file_name
            if not path.exists():
                summary["skipped_files"].append(file_name)
                continue
            digest = hashlib.sha256(path.read_bytes()).hexdigest()
            batch, created = ImportBatch.objects.get_or_create(
                file_hash=digest,
                defaults={"file_name": file_name, "source_type": source_type, "imported_by": user},
            )
            if not created and batch.status in {ImportBatch.Status.VALIDATED, ImportBatch.Status.POSTED}:
                continue
            batch.rows.all().delete()
            counts = self._read_xls(path, batch) if path.suffix.lower() == ".xls" else self._read_xlsx(path, batch)
            batch.status = ImportBatch.Status.VALIDATED
            batch.summary = counts
            batch.save(update_fields=["status", "summary", "updated_at"])
            summary["files"] += 1
            summary["sheets"] += counts["sheets"]
            summary["rows"] += counts["rows"]
        self.stdout.write(self.style.SUCCESS(f"Staging selesai: {summary}"))

    def _store(self, batch, sheet_name, row_number, values):
        cleaned = [canonical(value) for value in values]
        if not any(cleaned):
            return False
        status, errors = row_status(sheet_name, values)
        business_key = next((value.upper() for value in cleaned if re.fullmatch(r"(?:PO\d+|PJ\.FO\.[\w.]+)", value, re.I)), "")
        ImportRow.objects.create(
            batch=batch,
            sheet_name=sheet_name,
            row_number=row_number,
            business_key=business_key,
            fingerprint=fingerprint(cleaned),
            raw_payload={"values": cleaned},
            validation_errors=errors,
            status=status,
        )
        return True

    def _read_xlsx(self, path, batch):
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
        rows = 0
        for sheet in workbook.worksheets:
            for row_number, values in enumerate(sheet.iter_rows(values_only=True), 1):
                # Formatting in some invoice sheets expands max_column to 16k.
                trimmed = list(values[:80])
                while trimmed and trimmed[-1] is None:
                    trimmed.pop()
                rows += int(self._store(batch, sheet.title, row_number, trimmed))
        return {"sheets": len(workbook.sheetnames), "rows": rows}

    def _read_xls(self, path, batch):
        workbook = xlrd.open_workbook(path)
        rows = 0
        for sheet in workbook.sheets():
            for index in range(sheet.nrows):
                rows += int(self._store(batch, sheet.name, index + 1, sheet.row_values(index)[:80]))
        return {"sheets": workbook.nsheets, "rows": rows}
